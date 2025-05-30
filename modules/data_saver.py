import os
import re
import pandas as pd
from datetime import datetime
from .json_manager import JsonTermManager

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class DataSaver:
    def __init__(self):
        self.directories = {
            'csv': 'results/csv',
            'excel': 'results/excel'
        }
        self._create_directories()
        self.json_manager = JsonTermManager()

    def _create_directories(self):
        for directory in self.directories.values():
            if not os.path.exists(directory):
                os.makedirs(directory)

    def parse_ai_terms(self, ai_response):
        terms_data = []
        current_term = {}
        lines = [line.strip() for line in ai_response.split('\n') if line.strip()]
        number_pattern = r'^\d+\.'

        for line in lines:
            line = line.strip()
            is_numbered_term = bool(re.match(number_pattern, line) and ('Термин:' in line or '**Термин:**' in line))
            is_simple_term = line.startswith(('Термин:', '**Термин:**', '### Термин:'))

            if is_numbered_term or is_simple_term:
                if current_term:
                    terms_data.append(current_term)
                current_term = {}
                if '**Термин:**' in line:
                    term = line.split('**Термин:**')[1].strip()
                elif 'Термин:' in line:
                    term = line.split('Термин:')[1].strip()
                if is_numbered_term:
                    term = re.sub(number_pattern, '', term).strip()
                term = term.replace('#', '').strip()
                current_term['термин'] = term

            elif any(marker in line for marker in ['**Определение:**', 'Определение:', '- **Определение:**']):
                definition = line
                for marker in ['**Определение:**', 'Определение:', '- **Определение:**', '-']:
                    definition = definition.replace(marker, '')
                current_term['определение'] = definition.strip()

            elif any(marker in line for marker in ['**Перевод:**', 'Перевод:', '- **Перевод:**']):
                translation = line
                for marker in ['**Перевод:**', 'Перевод:', '- **Перевод:**', '-']:
                    translation = translation.replace(marker, '')
                current_term['перевод'] = translation.strip()

            elif any(marker in line for marker in ['**Релевантность:**', 'Релевантность:', '- **Релевантность:**']):
                relevance = line
                for marker in ['**Релевантность:**', 'Релевантность:', '- **Релевантность:**', '-']:
                    relevance = relevance.replace(marker, '')
                relevance = relevance.replace('%', '').strip()
                try:
                    current_term['релевантность'] = float(relevance)
                except ValueError:
                    current_term['релевантность'] = 0.0

        if current_term:
            terms_data.append(current_term)
        return terms_data

    def save_to_csv(self, terms_data, base_filename, metrics=None):
        if not terms_data:
            return False
        csv_path = os.path.join(self.directories['csv'], f"{base_filename}.csv")
        df = pd.DataFrame(terms_data)
        df.to_csv(csv_path, index=False, encoding='utf-8')
        if metrics:
            with open(csv_path, "a", encoding="utf-8") as f:
                f.write("\n")
                for k, v in metrics.items():
                    f.write(f"{k},{v}\n")
        print(f"Термины сохранены в CSV: {csv_path}")
        return True

    def save_to_excel(self, terms_data, base_filename, metrics=None):
        if not terms_data:
            return False
        excel_path = os.path.join(self.directories['excel'], f"{base_filename}.xlsx")
        df = pd.DataFrame(terms_data)
        df.to_excel(excel_path, index=False)
        if metrics:
            wb = load_workbook(excel_path)
            ws = wb.active
            start_row = ws.max_row + 2
            fill = PatternFill(start_color="FFF7D774", end_color="FFF7D774", fill_type="solid")
            for idx, (k, v) in enumerate(metrics.items()):
                ws.cell(row=start_row + idx, column=1, value=k).fill = fill
                ws.cell(row=start_row + idx, column=2, value=v).fill = fill
            wb.save(excel_path)
        print(f"Термины сохранены в Excel: {excel_path}")
        return True

    def prepare_terms_for_json(self, terms_data):
        try:
            json_terms = {}
            for term in terms_data:
                if isinstance(term, dict) and term.get('релевантность', 0) >= 80:
                    term_name = term.get('термин', '')
                    if term_name:
                        json_terms[term_name] = {
                            'term': term_name,
                            'definition': term.get('определение', ''),
                            'translation': term.get('перевод', ''),
                            'relevance': term.get('релевантность', 0),
                            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        }
            return json_terms
        except Exception as e:
            print(f"Ошибка при подготовке терминов для JSON: {str(e)}")
            return None

    def save_terms(self, ai_response, filename, metrics=None):
        terms_data = self.parse_ai_terms(ai_response)
        if not terms_data:
            print("В ответе ИИ не найдено терминов для сохранения")
            return False
        base_filename = os.path.splitext(filename)[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"terms_{timestamp}_{base_filename}"
        csv_saved = self.save_to_csv(terms_data, output_filename, metrics=metrics)
        excel_saved = self.save_to_excel(terms_data, output_filename, metrics=metrics)
        json_terms = self.prepare_terms_for_json(terms_data)
        added_count = 0
        updated_count = 0
        if json_terms:
            db_saved = self.json_manager.add_terms(json_terms)
            if db_saved:
                for term_data in json_terms.values():
                    if self.json_manager.term_exists(term_data['term']):
                        updated_count += 1
                    else:
                        added_count += 1
                print(f"\nСтатистика обработки терминов:")
                print(f"✨ Добавлено новых терминов: {added_count}")
                print(f"🔄 Обновлено существующих терминов: {updated_count}")
                print(f"📚 Всего обработано терминов: {added_count + updated_count}")
                print("Термины с релевантностью больше 80% успешно сохранены в базу данных")
            else:
                print("Произошла ошибка при сохранении терминов в базу данных")
        else:
            print("Нет терминов для сохранения в базу данных")
            db_saved = False
        return all([csv_saved, excel_saved, db_saved])
